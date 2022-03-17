# pip install openpyxl
from openpyxl import Workbook


### 엑셀 파일 생성, 데이터 작성하기
wb = Workbook()
# 새 워크북 생성
wb.save("sample.xlsx")
wb.close()


## sheet 제어
ws = wb.active
# 현재 활성화된 sheet 가져옴
ws.title = "sheet1"
# sheet1의 이름을 변경
ws2 = wb.create_sheet("Sheet2", 1)
# 1번째 index에 "Sheet2" 생성
ws2.sheet_properties.tabColor = "ff66ff"
# RGB 형태로 값을 넣어주면 탭 색상 변경

ws = wb["Sheet1"]
# Dict형태로 Sheet 명으로 접근
print(wb.sheetnames)
# 모든 Sheet 이름 확인

# Sheet 복사
copy1 = wb.copy_worksheet(ws)
copy1.title = "Copied Sheet1"


## 셀 입력
ws["A1"] = "Test"
# A1에 "test" 입력
print(ws["A1"])
# A1 셀의 정보를 출력
print(ws["A1"].value)
# A1 셀의 '값'을 출력

# row == 1, 2, 3, ...
# column == A(1), B(2), C(3), ...

print(ws.cell(column=2, row=1).value)
# ws["B1"].value

c = ws.cell(column=3, row=1, value=10)
# ws["C1"].value = 10
print(c.value)
# ws["C1"].value

# 반복문을 이용해서 랜덤 숫자 채우기
from random import *
index = 1
for x in range(1, 11): # 10개 row
    for y in range(1, 11): # 10개 column
        ws.cell(row=x, column=y, value=randint(0, 100)) # 0~100 사이의 숫자
        ws2.cell(row=x, column=y, value=index)
        index += 1

# 1줄씩 리스트 데이터 10줄 넣기
ws.append(["번호", "영어", "수학"]) # A, B, C
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])




### 엑셀 파일, 데이터 불러오기
from openpyxl import load_workbook

## 파일 불러오기
wb = load_workbook("sample.xlsx")
# sample.xlsx 파일에서 wb을 불러옴
ws = wb.active
# 활성화된 Sheet를 ws로 가져옴

# cell 데이터 불러오기
for x in range(1, 11):
     for y in range(1, 11):
         print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4 ..
     print()

# cell 갯수를 모를 때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4 ..
    print()

# 위와 똑같이 나오는지 
for row in ws.values:
    for cell in row:
        print(cell)

# 영어 column만 가지고 오기
col_B = ws["B"]
print(col_B)
for cell in col_B:
    print(cell.value)

# 영어, 수학 column 함께 가지고 오기
col_range = ws["B:C"]
for cols in col_range:
    for cell in cols:
        print(cell.value)

# 1번째 row만 가지고 오기
row_title = ws[1]
for cell in row_title:
    print(cell.value)

# 1번째 줄인 title 을 제외하고 2번째 줄에서 6번째 줄까지 가지고 오기
row_range = ws[2:6] 
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()


## 좌표정보(A/1)를 튜플(A,1) 형식으로 가져오기
from openpyxl.utils.cell import coordinate_from_string
row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
        print(cell.coordinate, end=" ") # 셀의 좌표정보(A/1) 가져오기
        xy = coordinate_from_string(cell.coordinate) # 좌표정보를 튜플(A,1) 형식으로 가져오기
        print(xy, end=" ")
        print(xy[0], end="") # A:튜플의 첫번째 정보
        print(xy[1], end=" ") # 1:튜플의 두번째 정보
    print()

# 전체 rows (tuple 사용)
print(tuple(ws.rows))
for row in tuple(ws.rows):
    print(row[2].value)

# 전체 columns (tuple 사용)
print(tuple(ws.columns))
for column in tuple(ws.columns):
    print(column[0].value)

# 전체 row (iter_rows() 사용)
for row in ws.iter_rows():
    print(row[2].value)

# 전체 column (iter_cols() 사용)
for column in ws.iter_cols():
    print(column[0].value)


## 범위를 지정하여 데이터 가져오기
# 2번째 줄부터 11번째 줄까지, 2번째 열부터 3번째 열까지
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
    print(row[0].value, row[1].value) # 수학, 영어
    print(row)
# 범위지정 디폴트는 1, Max

## 행렬 전환하여 데이터 가져오기
for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)



#### 엑셀 심화과정
### 엑셀 데이터 다루기
# if 사용
for row in ws.iter_rows(min_row=2):
    # 번호 row[0], 영어 row[1], 수학 row[2]
    if int(row[1].value) > 80:
        print(row[0].value, "번 학생은 영어 천재")

# 첫번째 셀의 영어를 컴퓨터로 수정하기
for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

## 행렬 삽입
ws.insert_rows(8)
# 8번째 줄에 1줄 추가
ws.insert_rows(8, 5)
# 8번째 줄에 5줄 추가
ws.insert_cols(2, 3)
# B번째 열에 3열 추가

## 행렬 삭제
ws.delete_rows(8)
# 8번째 줄 삭제
ws.delete_rows(8, 3)
# 8번째 줄부터 3줄 삭제
ws.delete_cols(2, 2)
# 2번재 열부터 2열 삭제

## 행렬 이동
ws.move_range("B1:C11", rows=0, cols=1)
ws["B1"].value = "국어" # B1 셀에 '국어' 입력
# 번호 영어 수학
# 번호 (국어) 영어 수학

## 차트
from openpyxl.chart import BarChart, Reference, LineChart
# B2:C11 까지의 데이터를 차트로 생성
bar_value = Reference(ws, min_row=2, max_row=11, min_col=2, max_col=3)
bar_chart = BarChart()         # 차트 종류 설정 (Bar, Line, Pie, ..)
bar_chart.add_data(bar_value)  # 차트 데이터 추가
ws.add_chart(bar_chart, "E1")  # 차트 넣을 위치 정의

# 제목을 포함한 B1:C11 까지의 데이터
line_value = Reference(ws, min_row=1, max_row=11, min_col=2, max_col=3)
line_chart = LineChart()
line_chart.add_data(line_value, titles_from_data=True) # 계열 -> 영어, 수학 (제목에서 가져옴)
line_chart.title = "성적표" # 제목
line_chart.style = 10 # 미리 정의된 스타일을 적용, 사용자가 개별 지정도 가능
line_chart.y_axis.title = "점수" # Y축의 제목
line_chart.x_axis.title = "번호" # X축의 제목
ws.add_chart(line_chart, "E1")

## 셀 스타일
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

# A열의 너비를 10으로 설정
ws.column_dimensions["A"].width = 10
# 1행의 높이를 10으로 설정
ws.row_dimensions[1].height = 10

# 셀 스타일 적용
a1 = ws["A1"] # 번호
b1 = ws["B1"] # 영어
c1 = ws["C1"] # 수학
a1.font = Font(color="FF0000", italic=True, bold=True) # 글자 색은 빨갛게, 이탤릭, 두껍게 적용
b1.font = Font(color="CC33FF", name="Arial", strike=True) # 폰트를 Arial 로 설정, 취소선 적용
c1.font = Font(color="0000FF", size=20, underline="single") # 글자 크기를 20으로, 밑줄 적용

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀은 초록색으로 적용
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # 각 cell에 대해서 정렬
        # center, left, right, top, bottom

        if cell.column == 1:
            continue
        # A열은 제외
        
        # cell이 정수형 데이터이고 90점보다 높으면
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") # 배경을 초록색으로 설정
            cell.font = Font(color="FF0000") # 폰트 색상 변경

# 틀 고정
ws.freeze_panes = "B2"
# B2기준으로 틀 고정

## 공식
import datetime
ws["A1"] = datetime.datetime.today() # 오늘 날짜 정보
ws["A2"] = "=SUM(1, 2, 3)" # 1 + 2 + 3 = 6 (합계)
ws["A3"] = "=AVERAGE(1, 2, 3)" # 2 (평균)

## 수식이 아닌 실제 데이터만 가져오기
wb = load_workbook("sample.xlsx", data_only=True)
# evaluate 되지 않은 상태의 데이터는 None 이라고 표시
for row in ws.values:
    for cell in row:
        print(cell)

## 셀 병합
ws.merge_cells("B2:D2") # B2 부터 D2 까지 합치겠음
ws["B2"].value = "Merged Cell"
ws.unmerge_cells("B2:D2")

## 이미지 삽입
from openpyxl.drawing.image import Image
img = Image("img.png")
ws.add_image(img, "C3")
# C3 위치에 img.png 파일의 이미지를 삽입
# ImportError : You must install Pillow to fetch image....
