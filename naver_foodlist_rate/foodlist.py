from selenium import webdriver
options = webdriver.ChromeOptions()
options.headless = True
options.add_argument("window-size=1920x1080")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.102 Safari/537.36")
browser = webdriver.Chrome('C:/github/project2/chromedriver.exe', options=options)

from selenium.webdriver.common.keys import Keys         # enter 입력키 가져오기
from selenium.webdriver.common.by import By # browser.find_element(By.ID, 'query') 이형식으로 바꾸기

import time, datetime

from bs4 import BeautifulSoup

import csv

import winsound # 사운드

start = time.time() # 시간재기용
# 엑셀 리스트 불러오기
from openpyxl import load_workbook # 엑셀파일 불러오기
wb = load_workbook('foodlist_all.xlsx') # sample.xlsx 파일에서 wb 을 불러옴
ws = wb.active # 활성화된 Sheet

si = []
for x in range(2, ws.max_row + 1):
    si.append(str(ws.cell(row=x, column=2).value))  # column=1는 A열

gu = []
for x in range(2, ws.max_row + 1):
    gu.append(str(ws.cell(row=x, column=3).value))  # column=2는 B열

end = time.time() # 시간재기용
sec = (end - start) # 시간재기용
sec2 = str(datetime.timedelta(seconds=sec)).split(".")
print(len(gu),"엑셀 리스트 읽기 완료",sec2[0])

# 엑셀 저장 파일 생성
filename = "food5-9.csv"
f = open(filename, "w", encoding="utf-8-sig", newline="")
writer = csv.writer(f)


# 페이지 이동
browser.get("https://map.naver.com/v5/search/")
time.sleep(3)

c = 1
for a in range(c-1, ws.max_row-1):
    elem = browser.find_element(By.CLASS_NAME, 'input_box').find_element(By.TAG_NAME, 'input')
    ## send_keys입력은 input 태그에서만 가능
    elem.clear()
    elem.send_keys(si[a], ' ', gu[a],' 맛집')
    elem.send_keys(Keys.ENTER)
    
    time.sleep(2)
    # iframe 전환
    browser.switch_to.frame('searchIframe') # frame ID 입력하여 frame 전환

    # 요즘뜨는[2] -> 리뷰많은[5] -> 많이찾는[1]
    elem = browser.find_element(By.XPATH, '//*[@id="app-root"]/div/div[1]/div/div/div/span[2]/a').click()
    elem = browser.find_element(By.XPATH, '//*[@id="app-root"]/div/div[1]/div[2]/div/ul/li[5]/a').click()
    time.sleep(2)


    end = time.time() # 시간재기용
    sec = (end - start) # 시간재기용
    sec2 = str(datetime.timedelta(seconds=sec)).split(".")
    print(si[a],gu[a],"시작 ",a+1,"/",ws.max_row-1," ",sec2[0])
    winsound.Beep(2000, 500) # winsound.Beep(frequency, duration)
    # frequency range:37 ~ 32767 , duration:ms

    for i in range(6):
        # 하나 엘리먼트 지정 -> 문서 높이를 가져와서 스크롤 무한 반복
        elem = browser.find_element(By.XPATH, '//*[@id="_pcmap_list_scroll_container"]')
        # 스크롤 무한반복
        prev_height = browser.execute_script('return arguments[0].scrollHeight', elem)
        while True:
            browser.execute_script('arguments[0].scrollBy(0, arguments[0].scrollHeight)', elem)
            time.sleep(1.5)
            cur_height = browser.execute_script('return arguments[0].scrollHeight', elem)
            if cur_height == prev_height:
                break
            prev_height = cur_height

        # 맛집 리스트 가져오기
        time.sleep(2)
        soup = BeautifulSoup(browser.page_source, "lxml")
        # foods = soup.find_all("div", attrs={"class":["ImZGtf mpg5gc", "Vpfmgd"]})
        foodname = soup.find_all("li", attrs={"class":"_1EKsQ _12tNp"})


        for food in foodname:
            # 음식점 이름 가져오기
            title = food.find("span", attrs={"class":"OXiLu"}).get_text()
            data = [si[a],gu[a],title]

            # 평점 및 리뷰 칸
            reviews = food.find_all("span", attrs={"class":"_2FqTn"})
            for review in reviews:
                review = review.get_text()
                data.append(review)

            # print(data)
            writer.writerow(data)

        print(i+1,"페이지 완료")
        # 다음 페이지 클릭
        elem = browser.find_element(By.XPATH, '//*[@id="app-root"]/div/div[2]/div[2]/a[7]').click()
        time.sleep(2)

    browser.switch_to.default_content() # 상위로 빠져 나옴

browser.quit()