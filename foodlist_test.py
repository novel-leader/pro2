from openpyxl import load_workbook # 파일 불러오기
import pyautogui
import pyperclip
import time
import sys

## 키보드로 자동화 끝내기 : Ctrl + Alt + Del
### pyautogui.mouseInfo() :마우스 정보를 알려주는 Tool 실행
### F1 누르면 X,Y,R,G,B,색 정보가 저장

# 이미지 인식 및 클릭 ------------------------------------------------
timeout = 10

def find_target(img_file, timeout=10):
    start = time.time()
    target = None
    while target is None:
        target = pyautogui.locateOnScreen(img_file, confidence=0.8) #"pip install opencv-python" 입력
        end = time.time()
        if end - start > timeout:
            break
    return target

def my_click(img_file, timeout=10):
    target = find_target(img_file, timeout)
    if target:
        pyautogui.click(target)
    else:
        print(f"[Timeout {timeout}s] Target not found ({img_file}). Terminate program.")
        sys.exit()
#--------------------------------------------------------------------    

# pyautogui 한글로 작성하기-------------------------------------------
def ko_write(text):
    pyperclip.copy(text) # pip install pyperclip
    pyautogui.hotkey("ctrl", "v")
# ko_write("함수로 한글 쓰기")
# pyautogui.write("영어 또는 숫자만 가능/한글 불가/값은 스트링만 가능")
#--------------------------------------------------------------------

# 엑셀 읽어오기-------------------------------------------------------
wb = load_workbook("foodlist.xlsx") # sample.xlsx 파일에서 wb 을 불러옴
ws = wb.active # 활성화된 Sheet

si = []
for x in range(2, ws.max_row + 1):
    si.append(str(ws.cell(row=x, column=2).value))  # column=1는 A열
print(si, len(si))

gu = []
for x in range(2, ws.max_row + 1):
    gu.append(str(ws.cell(row=x, column=3).value))  # column=2는 B열
print(gu, len(gu))

for i in range(0, ws.max_row-1):
    print(si[i], gu[i])
#---------------------------------------------------------------------

# 순서대로 읽기 -> 입력하기 -> 뒤로가기 : 반복3번
for i in range(0, ws.max_row-1):
    my_click("img1.png")
    pyautogui.sleep(1)
    ko_write(si[i]) # 값이 스트링일때만 입력가능 # 한글입력불가
    pyautogui.write(" ")
    ko_write(gu[i]) 
    ko_write(" 맛집")
    pyautogui.press("enter")
    pyautogui.sleep(3)

    pyautogui.hotkey("alt", "left")
    pyautogui.sleep(3)
